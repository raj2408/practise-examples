from django.shortcuts import render
from reg.models import *
from django.shortcuts import render_to_response
from django.http import HttpResponseRedirect
from django.http import HttpResponse

from django.contrib import auth
from reg.forms import *
import csv , io
from openpyxl import Workbook
import xlwt


def student_list(request):
    return render(request, 'student.html', {"objects": Student.objects.all()})

def teacher_list(request):
    return render(request, 'teacher.html', {"objects": Teacher.objects.all()})

def institute_list(request):
    return render(request, 'institute.html', {"objects":Institute.objects.all()})

def home(request):
    return render(request, 'home.html', {
        "length_rec": request.POST.get('length'),
        "breadth_rec": request.POST.get('breadth'),
    })

def signup():
    return render_to_response(request, 'signup.html', {})
    return render_to_response(request, 'signup.html', {})


def login(request):
    # c = {}
    # c.update(csrf(request))
    return render_to_response('login.html')

def auth_view(request):
    username = request.POST.get('username', '')
    password = request.POST.get('password','')
    user = auth.authenticate(username=username, password=password)

    if user is not None:
        auth.login(request,user)
        return HttpResponseRedirect('accounts/loggedin')
    else:
        return HttpResponseRedirect('/accounts/invalid')


def loggedin(request):
    return render_to_response('loggedin.html',{'full_name': request.user.username})

def invalid_login(request):
    return render_to_response('invalid_login.html')

def logout(request):
    auth.logout(request)
    return render_to_response('logout.html')

def createstudent(request):
    stu = StudentForm()
    if request.method == 'POST':
        stu = StudentForm(request.POST)
        if stu.is_valid():
            stu.save()
            return HttpResponseRedirect("/student")
    return render(request,"studentform.html",{'form':stu})

def editstudent(request,id):
    student=Student.objects.get(id=id)
    stu = StudentForm(instance=student)
    if request.method == 'POST':
        stu = StudentForm(request.POST, instance=student)
        if stu.is_valid():
            stu.save()
            return HttpResponseRedirect("/student")
    return render(request, "studentform.html", {'form': stu})


def deletestudent(request,id):
    student=Student.objects.get(id=id)
    student.delete()
    return HttpResponseRedirect("/student")

def createteacher(request):
    if request.method=='POST':
        form=TeacherForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data.get('name')
            email = form.cleaned_data.get('email')
            subject = form.cleaned_data.get('subject')
            Teacher.objects.create(name=name,email=email,subject=subject)
            #import pdb; pdb.set_trace()
            #Teacher.objects.create()
            return HttpResponseRedirect('/teacher')
    else :
        form=TeacherForm()

    return render(request, 'teacherform.html', {'form': form})

def createteacher_M(request):
    if request.method=='POST':
        form=TeacherFormM(request.POST)
        if form.is_valid():
            form.save()
            #import pdb; pdb.set_trace()
            #Teacher.objects.create()
            return HttpResponseRedirect('/teacher')
    else :
        form=TeacherFormM()

    return render(request, 'teacherformM.html', {'formM': form})

def editteacher(request,id):
    teacher = Teacher.objects.get(id = id)
    tchr = TeacherFormM(instance=teacher)
    if request.method == 'POST':
        tchr = TeacherFormM(request.POST, instance=teacher)
        if tchr.is_valid():
            tchr.save()
            return HttpResponseRedirect("/teacher")
    return render(request, "teacherform.html", {'form': tchr})

def createinstituteM(request):
    ins=InstituteFormM()
    if request.method == 'POST':
        ins = InstituteFormM(request.POST)
        if ins.is_valid():
            ins.save()
            return HttpResponseRedirect("/institute")
    return render(request,"instituteformM.html",{'form':ins})


#
def createinstituteC(request):
    if request.method=='POST':
        form=InstituteFormC(request.POST)
        if form.is_valid():
            name = form.cleaned_data.get('name')
            email = form.cleaned_data.get('email')
            website = form.cleaned_data.get('subject')
            Institute.objects.create(name=name,email=email,website=website)
            #import pdb; pdb.set_trace()
            #Teacher.objects.create()
            return HttpResponseRedirect('/institute')
    else :
        form = InstituteFormC()

    return render(request, "instituteformc.html",{'form': form})


def deleteinstitute(request,id):
    institute=Institute.objects.get(id=id)
    institute.delete()
    return HttpResponseRedirect("/institute")

def t_downloadcsv(request):
    items = Teacher.objects.all()
    response=HttpResponse(content_type = 'text/csv')
    response['Content-Disposition'] = 'attachment;filename=Teacher.csv'
    writer = csv.writer(response,delimiter=',')
    writer.writerow(['name','email','subject','institute'])
    for obj in items:
        writer.writerow([obj.name, obj.email, obj.subject, obj.institute])

    return response



def s_downloadxls(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Student.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Students')
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = ['Name', 'Website', 'Email', 'Institute', ]
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)     #what is this

    font_style = xlwt.XFStyle()

    rows = Student.objects.all().values_list('name', 'website', 'email', 'institute')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)     #whagt is this

    wb.save(response)
    return response


def i_downloadxlsx(request):
    instlist = Institute.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Institute.xlsx'
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Institute'

    # Define the titles for columns
    columns = [
        'Name',
        'Email',
        'Website',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for inst in instlist:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            inst.name,
            inst.email,
            inst.website,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


def login(request):
    return render(request, 'login.html')

def auth_view(request):
    username=request.POST.get('username','')
    password=request.POST.get('password','')
    user=auth.authenticate(username=username,password=password)

    if user is not None:
        auth.login(request,user)
        return HttpResponseRedirect('/loggedin')
    else:
        return HttpResponseRedirect('invalid_login')

def loggedin(request):
    return render(request, 'loggedin.html',{'name': request.user.username})





    return render(request, 'invalid_login.html')

def logout(request):
    auth.logout(request)
    return render(request, 'loggedout.html')







